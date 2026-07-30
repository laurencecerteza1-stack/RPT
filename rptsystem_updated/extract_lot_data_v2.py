#!/usr/bin/env python3
"""
extract_lot_data_v2.py — header-name-driven na bersyon (hindi na fixed
column offset). Gumagana kahit iba-iba ang bilang/pagkakasunod-sunod ng
mga column bawat subdivision file, basta may S/P/B/L + Status na header
row at may per-year block (year number sa isang row sa itaas ng header).

Gamit:
    python3 extract_lot_data_v2.py file1.xlsx file2.xlsx ... [-o outdir]

Lalabas: lot_details.csv at or_history.csv (pinagsama-samang lahat ng
files na pinasa) sa outdir (default: kasalukuyang folder).
"""
import sys, csv, re, os
import openpyxl

def norm(s):
    if s is None:
        return ''
    return re.sub(r'\s+', ' ', str(s)).strip().upper()

def find_header_row(ws, max_scan=15):
    for r, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1):
        vals = [norm(v) for v in row[:8]]
        if 'S' in vals and 'P' in vals and 'B' in vals and 'L' in vals:
            return r, row
    return None, None

def find_year_row(ws, header_row, max_scan_up=6):
    best = None
    for r in range(max(1, header_row - max_scan_up), header_row + 1):
        row = next(ws.iter_rows(min_row=r, max_row=r, values_only=True))
        hits = [(i, int(v)) for i, v in enumerate(row) if isinstance(v, (int, float)) and 2000 <= v <= 2035]
        if len(hits) >= 2:
            best = hits
    return best or []

def val(row, i):
    if i is None or i >= len(row):
        return ''
    v = row[i]
    if v is None:
        return ''
    if hasattr(v, 'strftime'):
        return v.strftime('%Y-%m-%d')
    return str(v).strip()

def find_col(header, names, start=0, end=None):
    end = len(header) if end is None else end
    for i in range(start, min(end, len(header))):
        h = norm(header[i])
        if not h:
            continue
        for n in names:
            if n in h:
                return i
    return None

# base field header keywords -> csv field
BASE_FIELDS = {
    'area': ['AREA'],
    'house_area': ['ITEM DESCRIPTION', 'ITEMDESCRIPTION'],
    'code': None,  # filled positionally after S/P/B/L
    'buyer': ['BUYER'],
    'location': ['LOCATION'],
    'lot_owner': ['LOT OWNER'],
    'title': ['TITLE'],
    'status': ['STATUS'],
    'remarks': ['REMARK'],
    'pin': ['PIN'],
}

def process_file(path, lot_rows, or_rows, seen):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    fname = os.path.basename(path)
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        header_row_n, header = find_header_row(ws)
        if not header_row_n:
            continue
        header = list(header)
        hnorm = [norm(h) for h in header]

        def idx_of(label):
            try:
                return hnorm.index(label)
            except ValueError:
                return None

        i_s, i_p, i_b, i_l = idx_of('S'), idx_of('P'), idx_of('B'), idx_of('L')
        i_area = find_col(header, BASE_FIELDS['area'])
        i_code = i_l + 2 if i_l is not None else None  # data col right after L (ItemDescription/code)
        i_buyer = find_col(header, BASE_FIELDS['buyer'])
        i_loc = find_col(header, BASE_FIELDS['location'])
        i_owner = find_col(header, BASE_FIELDS['lot_owner'])
        i_title = find_col(header, BASE_FIELDS['title'])
        i_status = find_col(header, BASE_FIELDS['status'])
        i_remarks = find_col(header, ['REMARK'])
        i_pin = find_col(header, ['PIN'])
        i_last_or = find_col(header, ['LAST OR'])
        i_soa_batch = find_col(header, ['SOA BATCH'])
        i_soa_year = find_col(header, ['SOA YR', 'SOA YEAR'])
        i_av = find_col(header, ['ASS. VALUE', 'ASSESSED VALUE', 'ASS VALUE', "ASS ("])
        # Hanapin LAHAT ng column na may "TD" sa header (hindi lang TD#1-3) —
        # kasi iba-iba ang bilang ng TD columns bawat subdivision file.
        td_cols = []
        for i, h in enumerate(hnorm):
            if not h:
                continue
            if re.search(r'\bTD\b', h) or 'TAX DEC' in h or 'TAX DECLARATION' in h:
                td_cols.append((i, header[i]))

        year_hits = find_year_row(ws, header_row_n)
        year_hits.sort(key=lambda x: x[0])
        blocks = []
        for k, (col, yr) in enumerate(year_hits):
            end = year_hits[k + 1][0] if k + 1 < len(year_hits) else len(header)
            blocks.append((yr, col, end))

        data_start = header_row_n + 1
        for row in ws.iter_rows(min_row=data_start, values_only=True):
            s_val = row[i_s] if i_s is not None and i_s < len(row) else None
            b_val = row[i_b] if i_b is not None and i_b < len(row) else None
            l_val = row[i_l] if i_l is not None and i_l < len(row) else None
            if not (s_val and b_val not in (None, '') and l_val not in (None, '')):
                continue
            sub = val(row, i_s)
            ph = val(row, i_p)
            blk = val(row, i_b)
            lot = val(row, i_l)
            if not sub or not blk or not lot:
                continue
            code = val(row, i_code)
            ra_number = ''  # walang RA/CTS column sa monitoring-type files na ito

            key = (fname, sheetname, sub, ph, blk, lot)
            if key not in seen:
                seen.add(key)
                td_vals = [(label, val(row, i)) for i, label in td_cols]
                td_vals = [(l, v) for l, v in td_vals if v]
                td1 = td_vals[0][1] if len(td_vals) > 0 else ''
                td2 = td_vals[1][1] if len(td_vals) > 1 else ''
                td3 = td_vals[2][1] if len(td_vals) > 2 else ''
                td_extra = ' | '.join(f'{l.strip()}: {v}' for l, v in td_vals[3:])

                lot_rows.append([
                    ra_number, sub, ph, blk, lot,
                    val(row, i_area), '', code, '',
                    val(row, i_buyer), val(row, i_loc), val(row, i_owner), val(row, i_title),
                    val(row, i_status), '', val(row, i_remarks),
                    val(row, i_pin),
                    td1, td2, td3, td_extra,
                    val(row, i_av), val(row, i_last_or),
                    val(row, i_soa_batch), val(row, i_soa_year),
                ])

            for yr, start, end in blocks:
                sub_header = header[start:end]
                i_as = find_col(sub_header, ['AS#', 'AS #'])
                i_jv = find_col(sub_header, ['JV#', 'JV #'])
                i_mc = find_col(sub_header, ['MC', 'LIAISON'])
                i_orno = find_col(sub_header, ['ACTUAL OR'])
                i_amt = find_col(sub_header, ['AMOUNT ON OR'])
                i_rmk = find_col(sub_header, ['REMARK'])

                def g(rel_i):
                    if rel_i is None:
                        return ''
                    return val(row, start + rel_i)

                or_no = g(i_orno)
                amt = g(i_amt)
                as_no = g(i_as)
                jv_no = g(i_jv)
                mc = g(i_mc)
                rmk = g(i_rmk)
                if not any([or_no, amt, as_no, jv_no, mc, rmk]):
                    continue
                or_rows.append([ra_number, sub, ph, blk, lot, yr, as_no, jv_no, mc, or_no, '', '', amt, '', rmk])

def main(argv):
    outdir = '.'
    if '-o' in argv:
        oi = argv.index('-o')
        outdir = argv[oi + 1]
        argv = argv[:oi] + argv[oi + 2:]
    files = argv
    os.makedirs(outdir, exist_ok=True)
    if not files:
        print("Gamit: python3 extract_lot_data_v2.py file1.xlsx [file2.xlsx ...] [-o outdir]")
        sys.exit(1)

    lot_rows, or_rows, seen = [], [], set()
    for f in files:
        print(f"Processing {f} ...")
        process_file(f, lot_rows, or_rows, seen)

    with open(os.path.join(outdir, 'lot_details.csv'), 'w', newline='', encoding='utf-8') as fh:
        w = csv.writer(fh)
        w.writerow(['ra_number','sub','ph','blk','lot','area','house_area','code','cts_no',
                     'buyer','location','lot_owner','title','status','date_fullpayment','remarks',
                     'pin','td1','td2','td3','td_extra','assessed_value','last_or',
                     'soa_batch','soa_year'])
        w.writerows(lot_rows)

    with open(os.path.join(outdir, 'or_history.csv'), 'w', newline='', encoding='utf-8') as fh:
        w = csv.writer(fh)
        w.writerow(['ra_number','sub','ph','blk','lot','year','as_no','jv_no','mc_liaison',
                     'or_number','fr','to','amount','date','remarks'])
        w.writerows(or_rows)

    print(f"lot_details.csv: {len(lot_rows)} lot(s)")
    print(f"or_history.csv:  {len(or_rows)} per-year OR entry(ies)")

if __name__ == '__main__':
    main(sys.argv[1:])