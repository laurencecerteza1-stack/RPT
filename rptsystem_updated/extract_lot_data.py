#!/usr/bin/env python3
"""
extract_lot_data.py — Kunin ang lot details + per-year OR history mula sa
ANUMANG subdivision Excel file na may parehong layout (S/P/B/L, Status,
per-year OR columns) — hindi CRC-specific, gagana ito sa ibang subdivisions
basta pareho ang column headers.

Paano gamitin:
    pip install openpyxl
    python3 extract_lot_data.py "SubdivisionName.xlsx"

Lalabas: lot_details.csv at or_history.csv sa parehong folder — i-upload
ang dalawang ito sa import_lot_full.php (Import Lot Details, tapos
Import OR History).

Inaasahang column headers (row na naglalaman ng "#","S","P","B","L", atbp
— hindi na kailangang fixed row number, hinahanap ito automatically):
  # | S | P | B | L | Area | House Area | Item Description | CTS No. |
  Buyer | Location | Lot Owner | Title | Status | Date of Fullpayment |
  REMARKS | PIN | TD #1 (...) | AV (...) | TD #3 (...) | AV (...) |
  Last OR# | SOA BATCH | SOA YEAR | ... tapos paulit-ulit na blocks
  ng 9 columns bawat taon: AS# | JV# | MC# / LIAISON | OR NUMBER | FR |
  TO | AMOUNT | DATE | REMARKS (may year label sa isang row sa itaas nito).
"""
import sys, csv, re
import openpyxl

def find_header_row(ws, max_scan=15):
    for r, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1):
        vals = [str(v).strip().upper() if v is not None else '' for v in row[:6]]
        if 'S' in vals and 'P' in vals and 'B' in vals and 'L' in vals:
            return r
    return None

def find_year_map(ws, max_scan=12):
    for row in ws.iter_rows(min_row=1, max_row=max_scan, values_only=True):
        hits = [(i, v) for i, v in enumerate(row) if isinstance(v, (int, float)) and 2000 <= v <= 2035]
        if len(hits) >= 3:
            return {int(v): i for i, v in hits}
    return {}

def val(row, i):
    if i is None or i >= len(row):
        return ''
    v = row[i]
    if v is None:
        return ''
    if hasattr(v, 'strftime'):
        return v.strftime('%Y-%m-%d')
    return str(v).strip()

def main(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    lot_rows, or_rows, seen = [], [], set()

    for name in wb.sheetnames:
        ws = wb[name]
        header_row = find_header_row(ws)
        if not header_row:
            continue
        ymap = find_year_map(ws)
        data_start = header_row + 1

        for row in ws.iter_rows(min_row=data_start, values_only=True):
            first = row[0] if len(row) > 0 else None
            s = row[1] if len(row) > 1 else None
            if not (isinstance(first, (int, float)) and s):
                continue
            sub, ph, blk, lot = val(row,1), val(row,2), val(row,3), val(row,4)
            if not sub or not blk or not lot:
                continue
            cts = val(row, 8)
            ra_number = cts if cts not in ('', '0') else ''

            key = (ra_number, sub, ph, blk, lot)
            if key not in seen:
                seen.add(key)
                td_current = val(row,19) or val(row,17)
                av = val(row,20) or val(row,18)
                lot_rows.append([
                    ra_number, sub, ph, blk, lot, val(row,5), val(row,6), val(row,7), cts,
                    val(row,9), val(row,10), val(row,11), val(row,12), val(row,13), val(row,14),
                    val(row,15), val(row,16), td_current, av, val(row,21), val(row,22), val(row,23),
                ])

            for year, col in ymap.items():
                block = [val(row, col + k) for k in range(9)]
                if not any(block):
                    continue
                or_rows.append([ra_number, sub, ph, blk, lot, year] + block)

    with open('lot_details.csv', 'w', newline='', encoding='utf-8') as f:
        w = csv.writer(f)
        w.writerow(['ra_number','sub','ph','blk','lot','area','house_area','code','cts_no',
                     'buyer','location','lot_owner','title','status','date_fullpayment','remarks',
                     'pin','td_current','assessed_value','last_or','soa_batch','soa_year'])
        w.writerows(lot_rows)

    with open('or_history.csv', 'w', newline='', encoding='utf-8') as f:
        w = csv.writer(f)
        w.writerow(['ra_number','sub','ph','blk','lot','year','as_no','jv_no','mc_liaison',
                     'or_number','fr','to','amount','date','remarks'])
        w.writerows(or_rows)

    print(f"lot_details.csv: {len(lot_rows)} lot(s)")
    print(f"or_history.csv:  {len(or_rows)} per-year OR entry(ies)")

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Gamit: python3 extract_lot_data.py \"Subdivision.xlsx\"")
        sys.exit(1)
    main(sys.argv[1])
